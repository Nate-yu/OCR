import os
import sys
os.environ['KMP_DUPLICATE_LIB_OK'] = 'True'
from PIL import Image
from pdf2image import convert_from_path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
import easyocr

excel_template_path = 'template.xlsx'
temp_img_path = "./temp_img/"

""" 
    用于将pdf文件夹中的文件逐个提取成为图片 
    pdf_folder_path: pdf文件夹的路径
    return: images，提取出来的图像列表
"""
def pdf2img(pdf_folder_path):
    # 初始化列表用于存储从pdf文件中提取出来的图片
    images = []
    # for循环遍历pdf文件夹中的每个pdf文件并将其转换为原始图像存入列表
    for pdf_file in sorted(os.listdir(pdf_folder_path)):
        if pdf_file.endswith('.pdf'):
            pdf_path = os.path.join(pdf_folder_path, pdf_file)
            image_list_pdf = convert_from_path(pdf_path)
            image = image_list_pdf[0]
            images.append(image)
    return images


""" 
    裁剪图片函数
    images: 图像列表
    return: person_info_img_cropped_list, 用户信息；test_data_img_cropped_list, 测试数据；data_curve_img_cropped_list, 数据曲线图
"""
def crop_images(images):
    # 初始化三个列表，用于存储裁剪后的图像（用户信息、测试数据以及数据曲线图）
    person_info_img_cropped_list = []
    test_data_img_cropped_list = []
    data_curve_img_cropped_list = []

    person_info_left = 245  # 用户信息左边界
    person_info_top = 360  # 用户信息上边界
    person_info_right = 1340  # 用户信息右边界
    person_info_bottom = 540  # 用户信息下边界

    test_data_left = 245  # 测试数据左边界
    test_data_top = 540  # 测试数据上边界
    test_data_right = 1050  # 测试数据右边界
    test_data_bottom = 1630  # 测试数据下边界

    data_curve_left = 1050  # 数据曲线图左边界
    data_curve_top = 540  # 数据曲线图上边界
    data_curve_right = 1340  # 数据曲线图右边界
    data_curve_bottom = 1630  # 数据曲线图下边界

    for i, image in enumerate(images):
        # 将图片保存以便进行OCR识别（也可以直接对PIL.Image对象进行处理）
        image.save(temp_img_path + f"page_{i + 1}.png")

        # 打开原图片开始进行裁剪
        img = Image.open(temp_img_path + f"page_{i + 1}.png")

        # 确保裁剪区域的坐标不会导致图像超出范围
        # 裁剪用户信息
        person_info_left = max(0, person_info_left)
        person_info_top = max(0, person_info_top)
        person_info_right = min(img.width, person_info_right)
        person_info_bottom = min(img.height, person_info_bottom)
        # 开始裁剪
        person_info_img_cropped = img.crop((person_info_left, person_info_top, person_info_right, person_info_bottom))
        person_info_img_cropped.save(temp_img_path + f"person_info_{i + 1}_cropped.png")
        person_info_img_cropped_list.append(person_info_img_cropped)

        # 裁剪测试数据
        test_data_left = max(0, test_data_left)
        test_data_top = max(0, test_data_top)
        test_data_right = min(img.width, test_data_right)
        test_data_bottom = min(img.height, test_data_bottom)
        # 开始裁剪
        test_data_img_cropped = img.crop((test_data_left, test_data_top, test_data_right, test_data_bottom))
        test_data_img_cropped.save(temp_img_path + f"test_data_{i + 1}_cropped.png")
        test_data_img_cropped_list.append(test_data_img_cropped)

        # 裁剪数据曲线图
        data_curve_left = max(0, data_curve_left)
        data_curve_top = max(0, data_curve_top)
        data_curve_right = min(img.width, data_curve_right)
        data_curve_bottom = min(img.height, data_curve_bottom)
        # 开始裁剪
        data_curve_img_cropped = img.crop((data_curve_left, data_curve_top, data_curve_right, data_curve_bottom))
        data_curve_img_cropped.save(temp_img_path + f"data_curve_{i + 1}_cropped.png")
        data_curve_img_cropped_list.append(data_curve_img_cropped)

    return person_info_img_cropped_list, test_data_img_cropped_list, data_curve_img_cropped_list


""" 
    使用Tesseract进行OCR识别 
    img_cropped_list: 裁剪后的图像列表
    return: person_info_list，识别的用户信息文本列表；test_data_list，识别的测试数据文本列表；test_data_full_list，识别的完整测试数据列表
"""
def ocr(person_info_img_cropped_list, test_data_img_cropped_list):
    # 用于存储两部分信息的列表
    person_info_list = []
    test_data_full_list = [] # 用于存储完整测试数据
    reader = easyocr.Reader(['ch_sim', 'en'])

    for i, person_info_img_cropped in enumerate(person_info_img_cropped_list):
        person_info = reader.readtext(f'./temp_img/person_info_{i + 1}_cropped.png', detail=0)
        person_info_list.append(person_info)

    for i, test_data_img_cropped in enumerate(test_data_img_cropped_list):
        test_full_data = reader.readtext(f'./temp_img/test_data_{i + 1}_cropped.png')
        test_data = [item[1] for item in test_full_data]
        # 用于保存测试日期，测试时间以及检查结论
        test_data_info = [] 
        test_data_info_list = []
        for i, item in enumerate(test_data):
            if item == '测试日期':
                test_data_info.append(test_data[i+1])
            elif item == '测试时间':
                test_data_info.append(test_data[i+1])
            if '意见' in item and i == len(test_data)-2:
                test_data_info.append(test_data[i+1])
            elif '意见' in item and i == len(test_data)-1:
                test_data_info.append(None)
        # test_data_list.append(test_data)
        test_data_info_list.append(test_data_info)
        for i in range(6):
            del test_full_data[0]
        test_data_full_list.append(test_full_data)
    return person_info_list, test_data_info_list, test_data_full_list


""" 
    提取基准元素坐标：提取前三个数字的锚框右上角的横坐标作为基准坐标
    data_list: 测试数据列表
    return: base_coordinate_list，基准元素坐标列表
"""
def extract_base_coordinate(data_list):
    base_coordinate = []
    base_coordinate_list = []
    count = 0
    for data in data_list:
        for item in data:
            value = None
            try:
                value = float(item[1])  # 尝试将元素转换为数值类型
            except ValueError:
                pass  # 若转换失败，则跳过该元素

            if value is not None:
                x_coord = item[0][1][0]  # 提取该数值元素坐标的第二个坐标的第一个值（横坐标）
                base_coordinate.append(x_coord)
                count += 1
                if count == 3:  # 达到三个数值元素后停止提取
                    break
        base_coordinate_list.append(base_coordinate)
    return base_coordinate_list

""" 
    预处理ocr识别出来的信息
    person_info_list: 用户信息列表
    test_full_data_list: 完整的带坐标的测试信息列表
    base_coordinate_list: 基准元素坐标列表
    return: person_info_dict_list, 用户信息字典列表；data_matrix_list, 测试数据矩阵列表
"""
def preprocess_text(person_info_list, test_full_data_list, base_coordinate_list):
    # 加载Excel工作簿和工作表
    template_excel_path = 'template.xlsx'
    # 加载Excel工作簿和工作表
    workbook = load_workbook(template_excel_path)
    sheet = workbook.active  # 假设数据填充在第一个工作表

    # 获取表头，只需要前11个元素作为个人信息的key
    headers = [cell.value for cell in next(sheet.iter_rows())]  # 假设表头在第一行
    headers = headers[:11]

    # 初始化一个字典来存储数据名称和数据值的配对
    person_info_dict_list = []
    person_info_dict = {header: None for header in headers}
    # 预处理患者个人信息：遍历数据列表，将数据名称和数据值配对
    index = 0
    for person_info in person_info_list:
        for i, item in enumerate(person_info):
            if item.endswith(":") or item.endswith("："):
                item = item[:-1]
            person_info[i] = item

        # 数据赋值给对应的键
        for header in person_info_dict.keys():
            if header in person_info:
                header_index = person_info.index(header)
                if header_index + 1 < len(person_info) and person_info[header_index + 1] not in person_info_dict.keys():
                    value = person_info[header_index + 1]
                    if header_index + 2 < len(person_info) and person_info[header_index + 2] not in person_info_dict:
                        value += '' + person_info[header_index + 2]
                        person_info_dict[header] = value
                if person_info[header_index + 1] in person_info_dict.keys():
                    continue
                person_info_dict[header] = value

        person_info_dict_list.append(person_info_dict)

    # 预处理测试数据
    data_matrix_list = []
    for base_coordinate in base_coordinate_list:
        # 初始化一个22x3的矩阵，用于暂时存储数据
        data_matrix = [[0 for _ in range(3)] for _ in range(22)]
        
        for test_full_data in test_full_data_list:
            i = 0
            for j, item in enumerate(test_full_data):
                # print(test_full_data[j])
                if (base_coordinate[0]-10 <= item[0][1][0] and item[0][1][0] <= base_coordinate[0]+20) and data_matrix[i][0] == 0:
                    data_matrix[i][0] = float(item[1])
                elif (base_coordinate[1]-10 <= item[0][1][0] and item[0][1][0] <= base_coordinate[1]+28) and data_matrix[i][1] == 0:
                    data_matrix[i][1] = float(item[1])
                    if data_matrix[i][0] == 0:
                        i += 1
                elif (base_coordinate[2]-10 <= item[0][1][0] and item[0][1][0] <= base_coordinate[2]+34) and data_matrix[i][2] == 0:
                    data_matrix[i][2] = float(item[1])
                    i += 1
        for row in data_matrix:
            if row[0] == 0:
                row[0]=None
            if row[1] == 0:
                row[1]=None
            if row[2] == 0:
                row[2]=None 
        data_matrix_list.append(data_matrix)
    
    return person_info_dict_list, data_matrix_list

""" 文本插入excel """
def text2excel(person_info_dict_list, data_matrix_list,test_data_info_list,data_curve_img_cropped_list): 
    # 加载Excel工作簿和工作表
    template_excel_path = 'template.xlsx'
    # 加载Excel工作簿和工作表
    workbook = load_workbook(template_excel_path)
    sheet = workbook.active  # 假设数据填充在第一个工作表

    # 居中对齐
    aligin_center = Alignment(horizontal='center', vertical='center')

    # 填充患者个人数据到Excel表格
    for i, person_info_dict in enumerate(person_info_dict_list,start=2):
        for j, (key, value) in enumerate(person_info_dict.items(), start=1):
            cell = sheet.cell(row=i, column=j, value=value)
            cell.alignment = aligin_center

    # 填充测试数据到Excel表格
    for k, test_data in enumerate(data_matrix_list, start=2):
        # 从第14列开始填充数据
        start_column = 14
        # 填充矩阵到第二行
        count = 0
        for i, row_data in enumerate(test_data):
            for j, cell_data in enumerate(row_data):
                cell = sheet.cell(row=k, column=start_column + count, value=cell_data)
                cell.alignment = aligin_center
                count += 1

    total_columns = sheet.max_column
    second_last_column = total_columns - 2  # 倒数第二列
    # 填充其他信息到Excel表格
    for i, test_data_info in enumerate(test_data_info_list, start=2):
        sheet.cell(row = i, column = 12, value = test_data_info[0]).alignment = aligin_center
        sheet.cell(row = i, column = 13, value = test_data_info[1]).alignment = aligin_center
        sheet.cell(row = i, column = second_last_column, value = test_data_info[2]).alignment = aligin_center

    # 填充数据曲线图到excel表格
    last_column = total_columns - 1 # 最后一列
    for i, data_curve_img_cropped in enumerate(data_curve_img_cropped_list, start=1):
        excel_image = ExcelImage(f'./temp_img/data_curve_{i}_cropped.png')
        sheet.column_dimensions[get_column_letter(last_column)].width = 40  # 不需要修改列宽，保持不变
        sheet.row_dimensions[i + 1].height = 380  # 将图片添加到下一行
        excel_image.height = 500  # 修改原图片高度
        sheet.add_image(excel_image, f"{get_column_letter(total_columns)}{i + 1}")  # 将图片添加到同一列的下一行

    # 保存工作簿
    workbook.save('template.xlsx')

""" 删除临时文件夹内的所有文件 """
def cleanup_temp_files(temp_folder_path):
    for filename in os.listdir(temp_folder_path):
        file_path = os.path.join(temp_folder_path, filename)
        try:
            if os.path.isfile(file_path):
                os.unlink(file_path)
            # 如果要删除子文件夹内的文件，可以使用递归
            # elif os.path.isdir(file_path): shutil.rmtree(file_path)
        except Exception as e:
            print(f"Failed to delete {file_path}. Reason: {e}")


def main(pdf_folder_path):
    images = []
    images = pdf2img(pdf_folder_path)
    person_info_img_cropped_list, test_data_img_cropped_list, data_curve_img_cropped_list = crop_images(images)
    person_info_list, test_data_info_list,test_full_data_list = ocr(person_info_img_cropped_list, test_data_img_cropped_list)
    base_coordinate_list = extract_base_coordinate(test_full_data_list) # 提取前三个数字坐标的第一个数据，用作后续数据的定位
    person_info_dict_list, data_matrix_list = preprocess_text(person_info_list, test_full_data_list, base_coordinate_list)
    text2excel(person_info_dict_list, data_matrix_list,test_data_info_list, data_curve_img_cropped_list)
    cleanup_temp_files(temp_img_path)

import tkinter as tk
from tkinter import filedialog
def select_folder():
    folder_path = filedialog.askdirectory()
    if folder_path:
        entry_path.delete(0, tk.END)
        entry_path.insert(0, folder_path)

def process_folder():
    folder_path = entry_path.get()
    if folder_path:
        try:
            # subprocess.run(["python", "OCR.py", folder_path])
            main(folder_path)
            tk.messagebox.showinfo("Success", "PDF 文件夹处理完成！")
        except Exception as e:
            tk.messagebox.showerror("Error", f"处理出错：{e}")
    else:
        tk.messagebox.showwarning("Warning", "请选择 PDF 文件夹路径！")

if __name__ == '__main__':
    # 创建主窗口
    root = tk.Tk()
    root.title("PDF 识别工具")
    window_width = 600
    window_height = 400
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x_coordinate = (screen_width - window_width) // 2
    y_coordinate = (screen_height - window_height) // 2
    root.geometry(f"{window_width}x{window_height}+{x_coordinate}+{y_coordinate}")

    # 添加按钮用于选择文件夹
    btn_browse = tk.Button(root, text="选择文件夹", command=select_folder)
    btn_browse.pack(pady=10)

    # 添加文本框用于显示选择的文件夹路径
    
    entry_path = tk.Entry(root, width=50)
    entry_path.pack(pady=5)

    # 添加按钮用于触发处理
    btn_process = tk.Button(root, text="处理 PDF 文件夹", command=process_folder)
    btn_process.pack(pady=5)

    # 运行主循环
    root.mainloop()
